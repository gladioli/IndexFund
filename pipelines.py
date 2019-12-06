# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: http://doc.scrapy.org/en/latest/topics/item-pipeline.html
import scrapy

from items import IndexFundItem
from openpyxl import Workbook
from openpyxl.worksheet import Worksheet
from openpyxl.cell import Cell
from openpyxl.styles import Font, colors, PatternFill, Side, Border
from openpyxl.styles.alignment import Alignment
from openpyxl.utils import get_column_letter


class IndexFundPipeline(object):
    wb: Workbook = None
    ws: Worksheet = None

    def __init__(self):
        self.titles = [
            "基金代码",
            "基金全称",
            "基金简称",
            "成立日期",
            "资产规模/亿元",
            "跟踪标的",
            "跟踪误差率",
            "基金公司",
            "基金公司规模/亿元",
            "管理费率/%(年)",
            "托管费率/%(年)",
        ]

        self.font_title = Font(name='Arial', size=14, color=colors.BLACK, bold=True)
        self.font_content = Font(name='Arial', size=12, color=colors.BLACK, bold=False)
        self.alignment_title = Alignment(horizontal='center', vertical='center')
        self.alignment_content = Alignment(horizontal='left', vertical='center')

    def open_spider(self, spider: scrapy.Spider):
        print('open_spider')
        self.wb = Workbook()
        self.wb.create_sheet('csi_500', 0)
        self.ws = self.wb.get_active_sheet()
        self.ws.append(self.titles)

    def close_spider(self, spider: scrapy.Spider):
        print('close_spider', self.ws.max_row, self.ws.max_column)
        for row in range(1, self.ws.max_row + 1):
            for col in range(1, self.ws.max_column + 1):
                try:
                    # print(col)
                    cell_index = '%s%d' % (self.num_to_letter(col), row)
                    print(cell_index)
                    self.ws[cell_index].font = Font(size=12 if row > 1 else 16, color=colors.BLACK, )
                    self.ws[cell_index].alignment = Alignment(horizontal='left' if row > 1 else 'center',
                                                              vertical='center', )
                except Exception as e:
                    print("Exception", e)

        self.wb.save('csi_500.xlsx')

    def process_item(self, item: IndexFundItem, spider: scrapy.Spider):
        row = [
            item['fund_code'] if 'fund_code' in item else '/',
            item['fund_fullname'] if 'fund_fullname' in item else '/',
            item['fund_name'] if 'fund_name' in item else '/',
            item['fund_release'] if 'fund_release' in item else '/',
            item['fund_size'] if 'fund_size' in item else '/',
            item['track_target'] if 'track_target' in item else '/',
            item['track_err_rate'] if 'track_err_rate' in item else '/',
            item['fund_glr'] if 'fund_glr' in item else '/',
            item['fund_gl_size'] if 'fund_gl_size' in item else '/',
            item['fund_gl_rate'] if 'fund_gl_rate' in item else '/',
            item['fund_tg_rate'] if 'fund_tg_rate' in item else '/',

        ]
        self.ws.append(row)
        return item

    def num_to_letter(self, c):
        s = ''
        while c > 0:
            m = c % 26
            m = 26 if m == 0 else m
            s += chr(m + 64)
            c = (c - m) // 26

        return s[::-1]
